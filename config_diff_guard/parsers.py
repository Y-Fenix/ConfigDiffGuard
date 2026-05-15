from __future__ import annotations

import csv
import io
import json
from pathlib import Path
from typing import Any

from .models import RowRecord, TableSnapshot
from .rules import RuleSet, infer_primary_key


SUPPORTED_EXTENSIONS = {".csv", ".tsv", ".json", ".xlsx"}


def parse_table(name: str, source: str, content: bytes, rules: RuleSet) -> TableSnapshot:
    ext = Path(name).suffix.lower()
    snapshot = TableSnapshot(name=name, source=source)
    try:
        if ext in {".csv", ".tsv"}:
            return _parse_delimited(name, source, content, rules, "\t" if ext == ".tsv" else None)
        if ext == ".json":
            return _parse_json(name, source, content, rules)
        if ext == ".xlsx":
            return _parse_xlsx(name, source, content, rules)
        snapshot.parse_errors.append(f"Unsupported extension: {ext}")
    except Exception as exc:  # noqa: BLE001 - report bad configs without crashing the whole run.
        snapshot.parse_errors.append(f"{type(exc).__name__}: {exc}")
    return snapshot


def _parse_delimited(
    name: str,
    source: str,
    content: bytes,
    rules: RuleSet,
    forced_delimiter: str | None,
) -> TableSnapshot:
    text = "\n".join(line for line in _decode(content).splitlines() if line.strip())
    sample = text[:4096]
    delimiter = forced_delimiter
    if delimiter is None:
        try:
            dialect = csv.Sniffer().sniff(sample)
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ","
    stream = io.StringIO(text)
    reader = csv.DictReader(stream, delimiter=delimiter)
    columns = list(reader.fieldnames or [])
    snapshot = TableSnapshot(name=name, source=source, columns=columns)
    key_fields = infer_primary_key(columns, rules.table_rule_for(name))
    for row_index, row in enumerate(reader, start=2):
        values = {str(k): _normalize(v) for k, v in row.items() if k is not None}
        if _is_empty_row(values):
            continue
        key = _build_key(values, key_fields, row_index)
        if key in snapshot.rows:
            snapshot.duplicate_keys[key] = snapshot.duplicate_keys.get(key, 1) + 1
            key = f"{key}#duplicate@{row_index}"
        snapshot.rows[key] = RowRecord(table=name, key=key, values=values, source=source, line=row_index)
    return snapshot


def _parse_json(name: str, source: str, content: bytes, rules: RuleSet) -> TableSnapshot:
    payload = json.loads(_decode(content))
    rows = _json_rows(payload)
    columns = sorted({key for row in rows for key in row})
    snapshot = TableSnapshot(name=name, source=source, columns=columns)
    key_fields = infer_primary_key(columns, rules.table_rule_for(name))
    for idx, row in enumerate(rows, start=1):
        values = {str(k): _normalize(v) for k, v in row.items()}
        if _is_empty_row(values):
            continue
        key = _build_key(values, key_fields, idx)
        if key in snapshot.rows:
            snapshot.duplicate_keys[key] = snapshot.duplicate_keys.get(key, 1) + 1
            key = f"{key}#duplicate@{idx}"
        snapshot.rows[key] = RowRecord(table=name, key=key, values=values, source=source, line=idx)
    return snapshot


def _parse_xlsx(name: str, source: str, content: bytes, rules: RuleSet) -> TableSnapshot:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        snapshot = TableSnapshot(name=name, source=source)
        snapshot.parse_errors.append("XLSX skipped: install openpyxl to parse Excel files.")
        return snapshot
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    columns: list[str] = []
    for sheet in workbook.worksheets:
        iterator = sheet.iter_rows(values_only=True)
        try:
            header = next(iterator)
        except StopIteration:
            continue
        local_columns = [str(cell).strip() if cell is not None else "" for cell in header]
        for row_number, values in enumerate(iterator, start=2):
            row = {
                col: value
                for col, value in zip(local_columns, values, strict=False)
                if col
            }
            row["__sheet"] = sheet.title
            row["__row"] = row_number
            rows.append(row)
            for col in row:
                if col not in columns:
                    columns.append(col)
    snapshot = TableSnapshot(name=name, source=source, columns=columns)
    key_fields = infer_primary_key(columns, rules.table_rule_for(name))
    for idx, row in enumerate(rows, start=1):
        values = {str(k): _normalize(v) for k, v in row.items()}
        if _is_empty_row(values):
            continue
        key = _build_key(values, key_fields, idx)
        if key in snapshot.rows:
            snapshot.duplicate_keys[key] = snapshot.duplicate_keys.get(key, 1) + 1
            key = f"{key}#duplicate@{idx}"
        snapshot.rows[key] = RowRecord(table=name, key=key, values=values, source=source, line=int(row.get("__row", idx)))
    return snapshot


def _json_rows(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        if all(isinstance(item, dict) for item in payload):
            return [dict(item) for item in payload]
        return [{"index": idx, "value": item} for idx, item in enumerate(payload)]
    if isinstance(payload, dict):
        for value in payload.values():
            if isinstance(value, list) and all(isinstance(item, dict) for item in value):
                return [dict(item) for item in value]
        return [_flatten_json(payload)]
    return [{"value": payload}]


def _flatten_json(value: Any, prefix: str = "") -> dict[str, Any]:
    if isinstance(value, dict):
        out: dict[str, Any] = {}
        for key, child in value.items():
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            out.update(_flatten_json(child, child_prefix))
        return out
    if isinstance(value, list):
        return {prefix: json.dumps(value, ensure_ascii=False, sort_keys=True)}
    return {prefix: value}


def _build_key(values: dict[str, Any], key_fields: list[str], row_index: int) -> str:
    if key_fields and all(field in values and values[field] not in ("", None) for field in key_fields):
        return "|".join(str(values[field]) for field in key_fields)
    return f"row:{row_index}"


def _normalize(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _is_empty_row(values: dict[str, Any]) -> bool:
    visible_values = [value for key, value in values.items() if not str(key).startswith("__") and str(key) != ""]
    return not visible_values or all(value in ("", None) for value in visible_values)


def _decode(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")
